from tkinter import *
from openpyxl import *
from tkinter import messagebox
wb=load_workbook('stud_info.xlsx')
sheet=wb.active
wb1=load_workbook('stud.xlsx')
sheet1=wb1.active


def create():
    def excel():
        sheet.column_dimensions['A'].width= 30
        sheet.column_dimensions['B'].width= 10
        sheet.column_dimensions['C'].width= 10
        sheet.column_dimensions['D'].width= 10
        sheet.column_dimensions['E'].width= 20
        sheet.column_dimensions['F'].width= 30
        sheet.column_dimensions['G'].width= 50

        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "Course"
        sheet.cell(row=1, column=3).value = "password"
        sheet.cell(row=1, column=4).value = "Sex"
        sheet.cell(row=1, column=5).value = "Contact no"
        sheet.cell(row=1, column=6).value = "Email id"
        sheet.cell(row=1, column=7).value = "Address"
    def focus1(event):
        course_field.focus_set()
    def focus2(event):
        password_field.focus_set()
    def focus3(event):
        sex_field.focus_set()
    def focus4(event):
        contact_no_field.focus_set()
    def focus5(event):
        email_id_field.focus_set()
    def focus6(event):
        address_field.focus_set()

    def clear():
        name.delete(0, END)
        course_field.delete(0, END)
        password_field.delete(0, END)
        sex_field.delete(0, END)
        contact_no_field.delete(0, END)
        email_id_field.delete(0, END)
        address_field.delete(0, END)
    def insert():
        if (name.get() == "" and
    		course_field.get() == "" and
    		password_field.get() == "" and
    		sex_field.get() == "" and
    		contact_no_field.get() == "" and
    		email_id_field.get() == "" and
    		address_field.get() == ""):

    		      print("empty input")

        else:
            current_row = sheet.max_row
            current_column = sheet.max_column

            sheet.cell(row=current_row + 1, column=1).value = name.get()
            sheet.cell(row=current_row + 1, column=2).value = course_field.get()
            sheet.cell(row=current_row + 1, column=3).value = password_field.get()
            sheet.cell(row=current_row + 1, column=4).value = sex_field.get()
            sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
            sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
            sheet.cell(row=current_row + 1, column=7).value = address_field.get()

            wb.save('stud_info.xlsx')
            name.focus_set()
            clear()
    root=Tk()
    root.configure(background='light green')
    root.title("student details")
    root.geometry("500x300")
    excel()

    head=Label(root, text="Form", bg="light green")
    head.grid(row=0, column=1)

    name=Label(root, text="Name", bg="light green")
    name.grid(row=1, column=0)

    course=Label(root, text="Course", bg="light green")
    course.grid(row=2, column=0)

    password=Label(root, text="password", bg="light green")
    password.grid(row=3, column=0)

    sex=Label(root, text="Sex", bg="light green")
    sex.grid(row=4, column=0)

    contact=Label(root, text="Contact No", bg="light green")
    contact.grid(row=5, column=0)

    mail=Label(root, text="E-mail id", bg="light green")
    mail.grid(row=6, column=0)

    add=Label(root, text="Address", bg="light green")
    add.grid(row=7, column=0)

    name = Entry(root)
    course_field = Entry(root)
    password_field = Entry(root)
    sex_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)

    name.bind("<Return>", focus1)
    course_field.bind("<Return>", focus2)
    password_field.bind("<Return>", focus3)
    sex_field.bind("<Return>", focus4)
    contact_no_field.bind("<Return>", focus5)
    email_id_field.bind("<Return>", focus6)

    name.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    password_field.grid(row=3, column=1, ipadx="100")
    sex_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")

    excel()

    submit=Button(root, text="Submit", fg="Black", bg="Red", command=insert)
    submit.grid(row=8, column=1)
    cancel=Button(root, text="CANCEL", command=root.destroy)
    cancel.grid(row=9, column=1)
    root.mainloop()

def check():
    m_row = sheet.max_row
    m1_row = sheet.min_row
    def clear2():
        ent.delete(0, END)
        ent1.delete(0, END)

    def store():
        def excel1():

            sheet1.column_dimensions['A'].width= 30
            sheet1.column_dimensions['B'].width= 10
            sheet1.column_dimensions['C'].width= 10

            sheet1.cell(row=1, column=1).value = "subject combination"
            sheet1.cell(row=1, column=2).value = "marks"
            sheet1.cell(row=1, column=3).value = "cgpa"

        def clear1():
            sub.delete(0, END)
            marks.delete(0, END)
            cgpa.delete(0, END)

        def insert1():
            if (sub.get() == "" and
        		marks.get() == "" and
        		cgpa.get() == ""):

        		      print("empty input")

            else:
                current_row = sheet.max_row
                current_column = sheet.max_column

                sheet1.cell(row=current_row + 1, column=1).value = sub.get()
                sheet1.cell(row=current_row + 1, column=2).value = marks.get()
                sheet1.cell(row=current_row + 1, column=3).value = cgpa.get()

                wb1.save('stud.xlsx')
                clear1()
        r=Tk()
        r.title("student details")
        r.geometry("500x300")
        excel1()
        subject=Label(r, text="subject combination")
        subject.grid(row=1, column=0)
        mark=Label(r, text="marks obtained")
        mark.grid(row=2, column=0)
        cgp=Label(r, text="cgpa / percentage")
        cgp.grid(row=3, column=0)
        sub=Entry(r)
        sub.grid(row=1, column=1, ipadx="100")
        marks=Entry(r)
        marks.grid(row=2, column=1, ipadx="100")
        cgpa=Entry(r)
        cgpa.grid(row=3, column=1, ipadx="100")
        excel1()
        submit=Button(r, text="Submit",command=insert1)
        submit.grid(row=5, column=1)
        cancel=Button(r, text="CANCEL", command=r.destroy)
        cancel.grid(row=6, column=1)
        r.mainloop()
    username = ent.get()
    password = ent1.get()

    for i in range(m1_row, m_row+1):
        cell_obj = sheet.cell(row =i , column = 1).value
        cell_pass = sheet.cell(row =i , column = 3).value
        if username == cell_obj and password == cell_pass:


            store()
        clear2()
    else:
        clear2()

        messagebox.showerror("Login error", "Incorrect username")

main=Tk()
main.title("student infomation")
main.geometry("500x300")
user=Label(main, text='USERNAME')
user.grid(row=1, column=0)
passw=Label(main, text='PASSWORD')
passw.grid(row=2, column=0)
ent=Entry(main)
ent.grid(row=1, column=1, ipadx='100')
ent1=Entry(main, show="*")
ent1.grid(row=2, column=1, ipadx='100')
b1=Button(main, text='SUBMIT', command=check)
b1.grid(row=3, column=1)
b3=Button(main, text='CREATE ACCOUNT', command=create)
b3.grid(row=4, column=1)
b2=Button(main, text='CANCEL', command=main.destroy)
b2.grid(row=5, column=1)
main.mainloop()
