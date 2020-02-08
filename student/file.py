import openpyxl


wb_obj = openpyxl.load_workbook("stud.xlsx")

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1).value
    cell = sheet_obj.cell(row = i, column = 3).value
    print("pass",cell)
    print(cell_obj)
