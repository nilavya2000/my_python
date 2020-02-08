from tkinter import *
from openpyxl import *

wb=load_workbook('data.xlsx')
sheet=wb.active

def excel():
    sheet.column_dimensions['A'].width= 30
    sheet.column_dimensions['B'].width= 10
    sheet.column_dimensions['C'].width= 10
    sheet.column_dimensions['D'].width= 10
    sheet.column_dimensions['E'].width= 20
    sheet.column_dimensions['F'].width= 30
    sheet.column_dimensions['G'].width= 50

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Sem"
    sheet.cell(row=1, column=4).value = "Sex"
    sheet.cell(row=1, column=5).value = "Contact no"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"

def focus1(event):
    course_field.focus_set()
def focus2(event):
    sem_field.focus_set()
def focus3(event):
    sex_field.focus_set()
def focus4(event):
    contact_no_field.focus_set()
def focus5(event):
    email_id_field.focus_set()
def focus6(event):
    address_field.focus_set()

def clear():
    name.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    sex_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

def insert():
    if (name.get() == "" and
		course_field.get() == "" and
		sem_field.get() == "" and
		sex_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""):

		      print("empty input")

    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = sex_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        wb.save('data.xlsx')
        name.focus_set()
        clear()


root=Tk()
root.configure(background='light green')
root.title("student details")
root.geometry("500x300")
excel()

head=Label(root, text="Form", bg="light green")
head.grid(row=0, column=1)

name=Label(root, text="Name", bg="light green")
name.grid(row=1, column=0)

course=Label(root, text="Course", bg="light green")
course.grid(row=2, column=0)

sem=Label(root, text="Sem", bg="light green")
sem.grid(row=3, column=0)

sex=Label(root, text="Sex", bg="light green")
sex.grid(row=4, column=0)

contact=Label(root, text="Contact No", bg="light green")
contact.grid(row=5, column=0)

mail=Label(root, text="E-mail id", bg="light green")
mail.grid(row=6, column=0)

add=Label(root, text="Address", bg="light green")
add.grid(row=7, column=0)

name = Entry(root)
course_field = Entry(root)
sem_field = Entry(root)
sex_field = Entry(root)
contact_no_field = Entry(root)
email_id_field = Entry(root)
address_field = Entry(root)

name.bind("<Return>", focus1)
course_field.bind("<Return>", focus2)
sem_field.bind("<Return>", focus3)
sex_field.bind("<Return>", focus4)
contact_no_field.bind("<Return>", focus5)
email_id_field.bind("<Return>", focus6)

name.grid(row=1, column=1, ipadx="100")
course_field.grid(row=2, column=1, ipadx="100")
sem_field.grid(row=3, column=1, ipadx="100")
sex_field.grid(row=4, column=1, ipadx="100")
contact_no_field.grid(row=5, column=1, ipadx="100")
email_id_field.grid(row=6, column=1, ipadx="100")
address_field.grid(row=7, column=1, ipadx="100")

excel()

submit=Button(root, text="Submit", fg="Black", bg="Red", command=insert)
submit.grid(row=8, column=1)
root.mainloop()
