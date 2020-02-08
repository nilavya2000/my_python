from tkinter import *
from tkinter import messagebox
from openpyxl import *
#import xlrd
#wb = xlrd.open_workbook('stud_info.xlsx')

wb=load_workbook('stud_info.xlsx')
sheet=wb.active
wb1=load_workbook('stud.xlsx')
sheet1=wb1.active

max_row = sheet.max_row

def excel():
    sheet.column_dimensions['A'].width= 30
    sheet.column_dimensions['B'].width= 10
    sheet.column_dimensions['C'].width= 10
    sheet.column_dimensions['D'].width= 10
    sheet.column_dimensions['E'].width= 20
    sheet.column_dimensions['F'].width= 30
    sheet.column_dimensions['G'].width= 50

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "password"
    sheet.cell(row=1, column=4).value = "Sex"
    sheet.cell(row=1, column=5).value = "Contact no"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"

def excel1():
    sheet1.column_dimensions['A'].width= 40
    sheet1.column_dimensions['B'].width= 40
    sheet1.column_dimensions['C'].width= 20

    sheet1.cell(row=1, column=1).value = "Subject combination"
    sheet1.cell(row=1, column=2).value = "marks obtained"
    sheet1.cell(row=1, column=3).value = "cgpa / percentage "

def clear1():
    nm.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    sex_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)
def clear():
    marks_field.delete(0, END)
    subject_field.delete(0, END)
    cgpa_field.delete(0, END)

def insert():
    if (nm.get() == "" and
		course_field.get() == "" and
		password_field.get() == "" and
		sex_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""):

		      print("empty input")

    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = nm.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = password_field.get()
        sheet.cell(row=current_row + 1, column=4).value = sex_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        wb.save('stud_info.xlsx')

        clear1()

def insert1():
    if (marks_field.get() == "" and
		subject_field.get() == "" and
		cgpa_field.get() == ""):

		      print("empty input")

    else:
        current_row = sheet1.max_row
        current_column = sheet1.max_column

        sheet1.cell(row=current_row + 1, column=1).value = Subject_field.get()
        sheet1.cell(row=current_row + 1, column=2).value = subject_field.get()
        sheet1.cell(row=current_row + 1, column=3).value = cgpa_field.get()

        wb1.save('stud.xlsx')
        clear()

def sub():
    main=Tk()
    main.geometry("500x300")
    main.title("WELCOME")
    def focus1(event):
        marks_field.focus_set()
    def focus2(event):
        subject_field.focus_set()
    def focus3(event):
        cgpa_field.focus_set()
    def focus4(event):
        e4.focus_set()

    fm=Label(main, text="subject combination")
    fm.grid(row=1, column=0)
    to=Label(main, text="marks obtained")
    to.grid(row=2, column=0)
    sub=Label(main, text="cgpa / percentage")
    sub.grid(row=3, column=0)
    #bd=Label(main, text="BODY")
    #bd.grid(row=4, column=0)
    marks_field=Entry(main)
    marks_field.grid(row=1, column=1, ipadx="100")
    subject_field=Entry(main)
    subject_field.grid(row=2, column=1, ipadx="100")
    cgpa_field=Entry(main)
    cgpa_field.grid(row=3, column=1, ipadx="100")
    #e4=Entry(main)
    #e4.grid(row=4, column=1, ipadx="100")
    sub=Button(main, text="SUBMIT", command=insert1)
    sub.grid(row=5, column=0)
    can=Button(main, text="CANCEL", command=main.destroy)
    can.grid(row=5, column=1)
    main.mainloop()

def submit():

    username = ent.get()
    password = ent1.get()

    for i in range(1, max_row + 1):
        cell_obj = sheet1.cell(row =i , column = 1)
        cell_pass = sheet1.cell(row =i , column = 3)
        if username == cell_obj and password == cell_pass:
            sub()
        else:
            messagebox.showerror("Login error", "Incorrect username")

def crt():
    r=Tk()
    r.configure(background='light yellow')
    r.title("student details")
    r.geometry("500x300")


    head=Label(r, text="Form", bg="light yellow")
    head.grid(row=0, column=1)

    name=Label(r, text="Name", bg="light yellow")
    name.grid(row=1, column=0)

    course=Label(r, text="Course", bg="light yellow")
    course.grid(row=2, column=0)

    sem=Label(r, text="Password", bg="light yellow")
    sem.grid(row=3, column=0)

    sex=Label(r, text="Sex", bg="light yellow")
    sex.grid(row=4, column=0)

    contact=Label(r, text="Contact No", bg="light yellow")
    contact.grid(row=5, column=0)

    mail=Label(r, text="E-mail id", bg="light yellow")
    mail.grid(row=6, column=0)

    add=Label(r, text="Address", bg="light yellow")
    add.grid(row=7, column=0)

    nm=Entry(r)
    course_field=Entry(r)
    sem_field=Entry(r)
    sex_field=Entry(r)
    contact_no_field=Entry(r)
    email_id_field=Entry(r)
    address_field=Entry(r)

    nm.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    sex_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")

    submit=Button(r, text="Submit", fg="Black", bg="Red", command=insert)
    submit.grid(row=8, column=1)
    excel()
    r.mainloop()

root=Tk()
root.geometry("500x300")
root.title("Student infomation")

l1=Label(root, text="username")
l1.grid(row=1,column=0)
l2=Label(root, text="password")
l2.grid(row=2, column=0)
ent=Entry(root)
ent.grid(row=1, column=1, ipadx="100")

ent1=Entry(root, show='*')
ent1.grid(row=2, column=1, ipadx="100")

b=Button(root, text="SUBMIT", command=submit)
b.grid(row=3, column=1)
c=Button(root, text="CANCEL", command=root.destroy)
c.grid(row=4, column=1)

l=Button(root, text="CREATE ACCOUNT", command=crt)
l.grid(row=5, column=1)
excel1()
root.mainloop()
